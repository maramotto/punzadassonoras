#!/usr/bin/env python3
"""Construye la hoja de calculo de revision a partir del dataset unificado (T1-T5)."""
import json, os, re, unicodedata
from collections import Counter, defaultdict
from urllib.parse import quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FUENTE = 'Arial'
TIPOS_COMPRABLES = {'libro', 'ensayo', 'cuento', 'relato', 'poema', 'obra de teatro'}


def norm(s):
    s = unicodedata.normalize('NFKD', (s or '').lower())
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'[^a-z0-9]+', ' ', s).strip()


def clave(a, o):
    return f'{norm(a)}||{norm(o)}'


def confianza(i):
    if not i or i.get('sin_resultados') or i.get('error'):
        return 'sin datos'
    if i.get('coincidencia_titulo') and i.get('coincidencia_autor'):
        return 'verificado'
    return 'inferido'


eps = json.load(open(f'{BASE}/data/refs_all.json'))
cache = json.load(open(f'{BASE}/data/cache_openlibrary.json'))

FILAS = []
for e in eps:
    tags = ', '.join(e.get('tags', []))
    desc_n = norm(e['descripcion'])
    for r in e['refs']:
        info = cache.get(clave(r.get('autor', ''), r.get('obra', '')), {})
        eds = info.get('editoriales') or []
        obra_n = norm(r.get('obra', ''))
        literal = 'sí' if obra_n and obra_n in desc_n else ('' if not obra_n else 'no — título normalizado o inferido')
        comprable = r.get('obra') and r.get('tipo') in TIPOS_COMPRABLES
        FILAS.append({
            'Serie': e['serie'],
            'Temporada': e['temporada'] or '',
            'Episodio': e['codigo'],
            'Título del episodio': e['titulo'],
            'Fecha': e['fecha'],
            'Autor': r.get('autor', ''),
            'Obra': r.get('obra', ''),
            'Tipo': r.get('tipo', ''),
            'Tags del episodio': tags,
            'Editorial (mencionada en el episodio)': r.get('editorial_mencionada', ''),
            'Editorial (Open Library)': ' / '.join(eds[:3]),
            'Año 1ª edición': info.get('anio_primera_edicion') or '',
            'Traductor': '',
            'Confianza del enriquecimiento': confianza(info),
            'Título citado literalmente': literal,
            'Notas': r.get('notas', ''),
            'Fuente de la referencia': e['fuente_datos'],
            'Link YouTube': e['url_youtube'],
            'Link Spotify': e['url_spotify'],
            'Link audio (MP3)': e['url_audio'],
            'Link TodosTusLibros': ('https://www.todostuslibros.com/busquedas?keyword=' +
                                    quote_plus(f"{r.get('obra','')} {r.get('autor','')}".strip())) if comprable else '',
        })

COLS = list(FILAS[0].keys())

wb = Workbook()

# ---------- Hoja 1: Referencias ----------
ws = wb.active
ws.title = 'Referencias'
ws.append(COLS)
for f in FILAS:
    ws.append([f[c] for c in COLS])

# ---------- Hoja 2: Episodios ----------
ws2 = wb.create_sheet('Episodios')
COLS2 = ['Serie', 'Temporada', 'Episodio', 'Título', 'Fecha', 'Duración (min)', 'Nº de referencias',
         'Tags', 'Fuente de los datos', '¿Transcripción disponible?', 'Link YouTube', 'Link Spotify', 'Notas']
ws2.append(COLS2)
for e in eps:
    ws2.append([e['serie'], e['temporada'] or '', e['codigo'], e['titulo'], e['fecha'],
                round((e.get('duracion_s') or 0) / 60), len(e['refs']), ', '.join(e.get('tags', [])),
                e['fuente_datos'], 'sí' if e['tiene_transcripcion'] else 'no',
                e['url_youtube'], e['url_spotify'], e.get('notas_episodio', '')])

# ---------- Hoja 3: Autores ----------
cuenta, obras, temps = Counter(), defaultdict(set), defaultdict(set)
for f in FILAS:
    if f['Autor']:
        cuenta[f['Autor']] += 1
        if f['Obra']:
            obras[f['Autor']].add(f['Obra'])
        if f['Temporada']:
            temps[f['Autor']].add(str(f['Temporada']))
ws3 = wb.create_sheet('Autores')
ws3.append(['Autor', 'Nº de menciones', 'Nº de obras distintas', 'Temporadas', 'Obras'])
for a, n in cuenta.most_common():
    ws3.append([a, n, len(obras[a]), ', '.join(sorted(temps[a])), ' · '.join(sorted(obras[a]))])

# ---------- Hoja 4: Leyenda ----------
ws4 = wb.create_sheet('Leyenda')
LEYENDA = [
    ['Punzadas Sonoras — referencias culturales (temporadas 1 a 5)', ''],
    ['', ''],
    ['Qué es esto', 'Extracción automática de las obras citadas en el podcast. Cubre las cinco temporadas más los episodios especiales y Las Glosas.'],
    ['', ''],
    ['De dónde sale cada temporada', ''],
    ['T3, T4, T5, especiales y Glosas', 'De las descripciones de los 66 vídeos de la playlist pública de YouTube. Son descripciones largas, con bibliografía detallada.'],
    ['T1 y T2', 'No están en YouTube. Se reconstruyeron desde el feed RSS público del podcast (feeds.megaphone.fm/PMSL3601016455), el mismo que alimenta Spotify.'],
    ['T2 desde el episodio 2x09', 'Además del feed, se usó la newsletter de las autoras (punzadas.substack.com), que da títulos exactos y editoriales que la descripción omite.'],
    ['', ''],
    ['Numeración de T1 y T2', 'ATENCIÓN: es una reconstrucción mía, no oficial. El feed no trae número de temporada ni de episodio. Ordené por fecha y puse el corte de temporada en el parón de verano de 2022 (T1: feb–ago 2022, 12 episodios; T2: sep 2022–jul 2023, 24 episodios). Si las autoras numeran de otra forma, hay que corregirlo aquí.'],
    ['', ''],
    ['Columnas que debes revisar tú', ''],
    ['Traductor', 'Vacía a propósito. Open Library casi nunca la trae y depende de la edición concreta. Se rellena a mano o con una fuente específica de ediciones españolas.'],
    ['Editorial (Open Library)', 'Puede corresponder a una edición distinta de la que se comenta en el episodio. Contrástala con la columna «Editorial (mencionada en el episodio)» cuando exista.'],
    ['Año 1ª edición', 'Año de la primera edición registrada, no el de la edición española que manejan en el podcast.'],
    ['', ''],
    ['Valores de «Confianza del enriquecimiento»', ''],
    ['verificado', 'Título y autor coinciden con el registro de Open Library. Fiable.'],
    ['inferido', 'Hay un registro parecido pero no coincide del todo. Revísalo antes de publicarlo.'],
    ['sin datos', 'Open Library no devolvió nada. Habitual en ensayo español de editorial pequeña, artículos académicos y tesis.'],
    ['(vacío)', 'No se buscó: son autores mencionados sin obra concreta, películas, series, obras de arte y otros formatos que no están en un catálogo de libros.'],
    ['', ''],
    ['Título citado literalmente', 'Comprobación automática: «sí» significa que ese título aparece tal cual en la descripción del episodio. «no» marca títulos que he normalizado o inferido (por ejemplo, cuando solo dicen «la figura Rapto, de Barthes» o «las memorias de Nabokov»). Esas filas son las que conviene revisar primero.'],
    ['Link Spotify', 'Es el enlace al programa, no al episodio. Los enlaces por episodio requieren la API de Spotify con credenciales, que no tengo.'],
    ['Link audio (MP3)', 'Solo para T1 y T2: el archivo de audio directo que sirve el feed. Útil si más adelante se transcriben con Whisper.'],
    ['Link TodosTusLibros', 'Es una búsqueda, no una ficha concreta: lleva al buscador de la librería con título y autor. La web no responde a peticiones automáticas, así que no están verificados uno a uno.'],
    ['', ''],
    ['Huecos conocidos', ''],
    ['Sin bibliografía en la descripción', '3x06 (Contingencia), 3x07 (Luz) y el especial «Los cuatro elementos: Fuego».'],
    ['Sin subtítulos en YouTube', '5x02 (En defensa de lo breve).'],
    ['T1 y T2 sin transcripción', 'Ninguno de los 36 episodios tiene subtítulos: no están en YouTube. Solo se han extraído las referencias de los textos, no del audio.'],
    ['Newsletter incompleta', 'El feed de Substack solo sirve los 20 posts más recientes. Las cartas de T1 (que incluían bibliografía e imágenes, como la de «Desrealidad») ya no son accesibles por esa vía.'],
    ['Numeración duplicada', 'Las Glosas «Sobre la salud sexual» y «Rodoreda, un bosque» aparecen ambas como 1x05 en los títulos originales.'],
    ['Episodios no cubiertos', 'El feed tiene 117 episodios y aquí hay 102. Faltan unos 15 de T3-T5 que no están en la playlist de YouTube y que podrían añadirse desde el RSS con el mismo método.'],
]
for fila in LEYENDA:
    ws4.append(fila)

# ---------- Formato ----------
AZUL = PatternFill('solid', fgColor='1F3864')
for hoja in wb.worksheets:
    for row in hoja.iter_rows():
        for c in row:
            c.font = Font(name=FUENTE, size=10)
    hoja.freeze_panes = 'A2'

for hoja in (ws, ws2, ws3):
    for c in hoja[1]:
        c.font = Font(name=FUENTE, size=10, bold=True, color='FFFFFF')
        c.fill = AZUL
        c.alignment = Alignment(vertical='center', wrap_text=True)
    hoja.row_dimensions[1].height = 30

ws4['A1'].font = Font(name=FUENTE, size=13, bold=True)
for i, fila in enumerate(LEYENDA, start=1):
    if fila[0] and not fila[1]:
        ws4[f'A{i}'].font = Font(name=FUENTE, size=10, bold=True)

ANCHOS = {'Serie': 16, 'Temporada': 11, 'Episodio': 11, 'Título del episodio': 38, 'Título': 38,
          'Fecha': 12, 'Autor': 30, 'Obra': 46, 'Tipo': 16, 'Tags del episodio': 34, 'Tags': 34,
          'Editorial (mencionada en el episodio)': 22, 'Editorial (Open Library)': 30,
          'Año 1ª edición': 13, 'Traductor': 18, 'Confianza del enriquecimiento': 16,
          'Título citado literalmente': 22, 'Notas': 40, 'Fuente de la referencia': 22,
          'Link YouTube': 44, 'Link Spotify': 46, 'Link audio (MP3)': 40, 'Link TodosTusLibros': 46,
          'Duración (min)': 13, 'Nº de referencias': 13, 'Fuente de los datos': 22,
          '¿Transcripción disponible?': 14, 'Nº de menciones': 13, 'Nº de obras distintas': 13,
          'Temporadas': 13, 'Obras': 70}
for hoja in (ws, ws2, ws3):
    for i, c in enumerate(hoja[1], start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ANCHOS.get(c.value, 20)
ws4.column_dimensions['A'].width = 42
ws4.column_dimensions['B'].width = 110
for row in ws4.iter_rows(min_col=2, max_col=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical='top')

for hoja, nombre in ((ws, 'TablaRefs'), (ws2, 'TablaEps'), (ws3, 'TablaAutores')):
    ref = f'A1:{get_column_letter(hoja.max_column)}{hoja.max_row}'
    t = Table(displayName=nombre, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name='TableStyleLight9', showRowStripes=True)
    hoja.add_table(t)

salida = f'{BASE}/Punzadas_Sonoras_referencias.xlsx'
wb.save(salida)
print('filas de referencias:', len(FILAS))
print('episodios:', len(eps))
print('guardado en', salida)
